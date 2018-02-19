
import sys
import copy
import numpy as np

# # use one of the multiple public chemical databases to retrive InChi strings
# ChemSpider database
from chemspipy import ChemSpider, Compound
# PubChemPy database
import pubchempy as pcp

# to handle I/O with excel file
from openpyxl import Workbook, load_workbook

# connect to ChemSpider using your security token
cs = ChemSpider('security token goes here')


def read_compound_names_from_excel_file(file_name):
    """"""
    wb = load_workbook(file_name)
    # I assume we only need to consider the first sheet
    ws1 = wb.active
    # I assume the first column holds the names of the compounds
    # I assume the first row contains headers

    # apparently this doesn't work anymore? the dictionary is empty?
    # print(ws1.row_dimensions.keys()).
    # nRows = len(ws1.row_dimensions.keys())
    # names = np.array([str(cell[0].value) for cell in ws1.iter_rows(min_row=2, max_col=1, max_row=nRows-1)])

    # We assume that all rows are empty after the last compound
    names = np.array([str(cell[0].value) for cell in ws1.iter_rows(min_row=2, max_col=1)])
    return names


def create_compound_list(list_data):
    """"""
    # initialize the return list
    list_compound = np.array([None] * len(list_data))


    for index, name, in enumerate(list_data):
        tList = [compound for compound in cs.search(name)]
        # tList = [compound for compound in cs.get_compounds(name, 'name')] # for pcp
        if tList == []:
            print("Could not find compound {:s}".format(name))
            list_compound[index] = None
            continue
        elif len(tList) > 1:
            print("More than one compound of {:s} found".format(name))

        print("Using compound {:d} for {:s}".format(tList[0].csid, name))
        list_compound[index] = cs.get_compound(tList[0].csid)
        # list_compound[index] = tList[0] # for pcp

    return list_compound


def inchi_to_ascii(list_compound):
    """"""
    list_inchi = [compound.inchi if (compound is not None) else "\x00" for compound in list_compound]

    list_ascii = [ [ord(char) for char in inchi] for inchi in list_inchi]

    return list_ascii


def smiles_to_ascii(list_compound):
    """"""
    list_smiles = [compound.smiles if (compound is not None) else "\x00" for compound in list_compound]
    # list_smiles = [compound.isomeric_smiles if (compound is not None) else "\x00" for compound in list_compound] # for pcp

    list_ascii = [ [ord(char) for char in smiles] for smiles in list_smiles]

    return list_ascii


def save_to_excel_file_by_rows(compound_names, list_representation, file_name):
    """"""
    wb = Workbook()
    ws1 = wb.active

    vectors = copy.deepcopy(list_representation)

    # find the max length of the vectors
    maxLen = max([len(vec) for vec in vectors])

    # insert compound names
    for idx, vec in enumerate(vectors):
        vec.insert(0, compound_names[idx])

    # header
    vectors.insert(0, ["Compound Name", ])

    # write the actual data
    for vec in vectors:
        ws1.append(vec)

    wb.save(file_name)
    return


def save_to_excel_file_by_cols(compound_names, list_representation, file_name):
    """"""
    wb = Workbook()
    ws1 = wb.active

    vectors = copy.deepcopy(list_representation)

    # find the max length of the vectors
    maxLen = max([len(vec) for vec in vectors])

    for idx, vec in enumerate(vectors):
        diffLen = maxLen - len(vec)
        # fill empty slots with zeros
        vec.extend([0]*diffLen)
        assert(len(vec) == maxLen)
        # insert compound name
        vec.insert(0, compound_names[idx])

    # the worksheet is blank so we select what rows + columns we will fill
    # use maxLen+1 since one of the rows is the name of the compound
    column_iterator = ws1.iter_cols(min_row=1, max_row=maxLen+1, min_col=1, max_col=len(vectors))

    # fill the worksheet
    for iCol, col in enumerate(column_iterator):
        for iCell, cell in enumerate(col):
            cell.value = vectors[iCol][iCell]

    wb.save(file_name)
    return


def convert_to_inchi_representation(file_name):
    """"""
    list_name = read_compound_names_from_excel_file(file_name)
    list_compound = create_compound_list(list_name)
    list_ascii = inchi_to_ascii(list_compound)
    save_to_excel_file_by_cols(list_name, list_ascii, file_name.replace(".xlsx", "_ascii.xlsx"))
    return


def convert_to_smiles_representation(file_name):
    """"""
    list_name = read_compound_names_from_excel_file(file_name)
    list_compound = create_compound_list(list_name)
    list_smiles = smiles_to_ascii(list_compound)
    save_to_excel_file_by_cols(list_name, list_smiles, file_name.replace(".xlsx", "_smiles.xlsx"))
    return


if __name__ == "__main__":
    """"""
    convert_to_inchi_representation('./example_data_set.xlsx')
    convert_to_smiles_representation('./example_data_set.xlsx')